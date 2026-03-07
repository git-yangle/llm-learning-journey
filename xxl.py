"""将 line.csv 转换为带样式的 Excel 和 PDF 文件"""
import csv
from pathlib import Path

BASE_DIR = Path(__file__).parent
CSV_PATH  = BASE_DIR / "line.csv"
XLSX_PATH = BASE_DIR / "line.xlsx"
PDF_PATH  = BASE_DIR / "line.pdf"

# 读取 TSV
with open(CSV_PATH, encoding="utf-8") as f:
    rows = [row for row in csv.reader(f, delimiter="\t") if any(row)]

# 各方案颜色（与地图路线一致）
ROUTE_COLORS = {1: "E74C3C", 2: "2980B9", 3: "27AE60", 4: "8E44AD"}

# ── Excel ──────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "行程对比"

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for r_idx, row in enumerate(rows, start=1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.alignment = Alignment(
            wrap_text=True, vertical="center",
            horizontal="center" if c_idx == 1 else "left",
        )
        cell.border = border

        is_header  = r_idx == 1
        is_total   = r_idx == len(rows)
        is_day_col = c_idx == 1

        if is_header:
            bg    = ROUTE_COLORS.get(c_idx - 1, "2C3E50") if c_idx > 1 else "2C3E50"
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif is_total:
            cell.fill = PatternFill("solid", fgColor="ECF0F1")
            cell.font = Font(bold=True, size=9, color="2C3E50")
        elif is_day_col:
            cell.fill = PatternFill("solid", fgColor="2C3E50")
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif "自由" in str(value) or value.strip() == "0km":
            cell.fill = PatternFill("solid", fgColor="F8F9FA")
            cell.font = Font(color="AAAAAA", size=9)
        else:
            bg = "FFFFFF" if r_idx % 2 == 0 else "FDFEFE"
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=9)

# 列宽 & 行高
ws.column_dimensions["A"].width = 10
for i in range(2, len(rows[0]) + 1):
    ws.column_dimensions[get_column_letter(i)].width = 32
ws.row_dimensions[1].height = 36
for i in range(2, len(rows) + 1):
    ws.row_dimensions[i].height = 42
ws.freeze_panes = "B2"

wb.save(XLSX_PATH)
print(f"Excel 已生成：{XLSX_PATH}")

# ── PDF ────────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

# 注册内置中文 CID 字体
pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
CN = "STSong-Light"

PAGE_W, PAGE_H = landscape(A4)
doc = SimpleDocTemplate(
    str(PDF_PATH),
    pagesize=landscape(A4),
    leftMargin=15 * mm, rightMargin=15 * mm,
    topMargin=12 * mm, bottomMargin=12 * mm,
)

title_style = ParagraphStyle(
    "title", fontName=CN, fontSize=14, alignment=1,
    spaceAfter=6, textColor=colors.HexColor("#2C3E50"),
)
cell_style = ParagraphStyle(
    "cell", fontName=CN, fontSize=7.5, leading=11, wordWrap="CJK",
)
header_style = ParagraphStyle(
    "hdr", fontName=CN, fontSize=8.5, textColor=colors.white,
    leading=12, alignment=1,
)

def p(text, style=None):
    return Paragraph(str(text), style or cell_style)

def ph(text):
    return Paragraph(str(text), header_style)

n_cols = len(rows[0])
# 首列固定宽，其余等分剩余空间
avail  = PAGE_W - 30 * mm
col_widths = [18 * mm] + [(avail - 18 * mm) / (n_cols - 1)] * (n_cols - 1)

table_data = []
for r_idx, row in enumerate(rows):
    if r_idx == 0:
        table_data.append([ph(v) for v in row])
    else:
        table_data.append([p(v) for v in row])

n_rows = len(table_data)
DARK  = colors.HexColor("#2C3E50")
LGRAY = colors.HexColor("#ECF0F1")
WHITE = colors.white

ROUTE_HEX = {
    1: colors.HexColor("#E74C3C"),
    2: colors.HexColor("#2980B9"),
    3: colors.HexColor("#27AE60"),
    4: colors.HexColor("#8E44AD"),
}

ts = TableStyle([
    ("FONTNAME",      (0, 0), (-1, -1), CN),
    ("FONTSIZE",      (0, 0), (-1, -1), 8),
    ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
    ("ROWBACKGROUNDS", (1, 1), (-1, n_rows - 2),
     [WHITE, colors.HexColor("#FDFEFE")]),
    # 表头行
    ("BACKGROUND",    (0, 0), (0, 0), DARK),
    ("TEXTCOLOR",     (0, 0), (-1, 0), WHITE),
    ("FONTSIZE",      (0, 0), (-1, 0), 9),
    ("ALIGN",         (0, 0), (-1, 0), "CENTER"),
    # 各方案列颜色
    *[("BACKGROUND",  (i, 0), (i, 0), ROUTE_HEX[i]) for i in range(1, n_cols)],
    # 天数列
    ("BACKGROUND",    (0, 1), (0, n_rows - 2), DARK),
    ("TEXTCOLOR",     (0, 1), (0, n_rows - 2), WHITE),
    ("ALIGN",         (0, 0), (0, -1), "CENTER"),
    # 总计行
    ("BACKGROUND",    (0, n_rows - 1), (-1, n_rows - 1), LGRAY),
    ("FONTSIZE",      (0, n_rows - 1), (-1, n_rows - 1), 8),
    # 内边距
    ("TOPPADDING",    (0, 0), (-1, -1), 5),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
])

table = Table(table_data, colWidths=col_widths, repeatRows=1)
table.setStyle(ts)

doc.build([
    Paragraph("新西兰南岛自驾路线对比", title_style),
    Spacer(1, 4 * mm),
    table,
])
print(f"PDF  已生成：{PDF_PATH}")
